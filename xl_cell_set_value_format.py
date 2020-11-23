import openpyxl
import datetime

wb = openpyxl.load_workbook("売上データ.xlsx")
ws = wb["4月"]

c1 = ws["A10"]
c1.value = datetime.datetime(2020, 4, 5)
c1.number_format = "yyyy/m/d"

c2 = ws["B10"]
c2.value = "株式会社 鈴木商店"

c3 = ws["C10"]
c3.value = "商品Z"

c4 = ws["D10"]
c4.value = 3800
c4.number_format = "#,##0"

c5 = ws["E10"]
c5.value = 12

c6 = ws["F10"]
c6.value = "=D10 * E10"
c6.number_format = "#,##0"

wb.save("売上データ.xlsx")