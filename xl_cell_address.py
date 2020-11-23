import openpyxl

wb = openpyxl.load_workbook("売上データ.xlsx")
ws = wb["4月"]

# セルの番地指定
c = ws["A1"]
print(c.coordinate)
print(c.row)
print(c.column)

# セルの行列番号指定
c2 = ws.cell(1, 1)
print(c2.coordinate)
print(c2.row)
print(c2.column)