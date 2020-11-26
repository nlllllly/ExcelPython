import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook("売上データ.xlsx")
ws = wb["6月"]

# 書き込むデータ
data_list = [
    [datetime(2020, 6, 1), "株式会社 ZZ社", "商品Z", 7200, 30],
    [datetime(2020, 6, 3), "株式会社 wssx社", "商品aaa", 7200, 15],
    [datetime(2020, 6, 9), "株式会社 JJJ社", "商品J", 1200, 20]
]

# 書き込み開始row番号
row_num = 4

for row in data_list:
    c1 = ws["A" + str(row_num)]
    c1.value = row[0]
    c1.number_format = "yyyy/m/d"
    
    c2 = ws["B" + str(row_num)]
    c2.value = row[1]
    
    c3 = ws["C" + str(row_num)]
    c3.value = row[2]

    c4 = ws["D" + str(row_num)]
    c4.value = row[3]

    c5 = ws["E" + str(row_num)]
    c5.value = row[4]

    c6 = ws["F" + str(row_num)]
    # セル数式の組み立て
    c6.value = "=D" + str(row_num) + "*E" + str(row_num)

    # 次のセルに移る
    row_num = row_num + 1


wb.save("売上データ.xlsx")