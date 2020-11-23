import openpyxl

wb = openpyxl.load_workbook("売上データ.xlsx")
ws = wb["4月"]

c1 = ws["A4"]
print(c1.value)
c2 = ws["B4"]
print(c2.value)
c3 = ws["C4"]
print(c3.value)
c2 = ws["D4"]
print(c2.value)
c5 = ws["E4"]
print(c5.value)
c6 = ws["F4"]
print(c6.value)
